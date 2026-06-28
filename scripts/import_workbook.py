#!/usr/bin/env python3
"""Parse the client's garment workbook into seed.json for the ERP demo.

Source of truth = CONSOLIDATED STATUS (flat, 250 rows) + VENDOR_LIST.
We normalize the real-world mess (units, junk text in number cells, vendor
combos, missing SKUs) and synthesize the bits the workbook never tracked
(opening fabric stock, size breakups, multi-event dispatch logs).
"""
import json, os, re, datetime
import openpyxl

_WB_NAME = "DAILY UPDATES STOCK FORM 28-11-2025 (1).xlsx"
_WB_CANDIDATES = [
    os.path.expanduser(f"~/Downloads/Sportsun Factory Data/{_WB_NAME}"),
    os.path.expanduser(f"~/Downloads/{_WB_NAME}"),
]
WB = next((p for p in _WB_CANDIDATES if os.path.exists(p)), _WB_CANDIDATES[0])
OUT = os.path.join(os.path.dirname(__file__), "..", "prisma", "seed.json")

# CONSOLIDATED STATUS column indices
C_ORDER_DATE, C_SKU, C_MRP, C_STYLE, C_ITEM, C_CUT, C_DISP = 1, 2, 3, 4, 5, 6, 7
C_FABRIC, C_FAB_ISSUE_DATE, C_AVG, C_UNIT = 9, 10, 11, 12
C_FAB_ISSUED, C_CONSUMED, C_CUTMASTER, C_VENDOR = 13, 14, 16, 17
C_CUT_ISSUED_ON, C_ETD, C_REMARK, C_STATUS = 18, 19, 20, 21

JUNK = ("TBC", "PENDING", "COLORWISE", "NOT RECVD", "NOT RECEIVED", "?")
SIZE_RATIO = [("S", .08), ("M", .17), ("L", .25), ("XL", .25), ("2XL", .17), ("3XL", .08)]


def num(v):
    """Coerce a cell to a float, or None if it's junk/text/empty."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().upper()
    if not s or any(j in s for j in JUNK):
        return None
    s = s.replace("/-", "").replace(",", "").replace("₹", "")
    m = re.search(r"-?\d+(\.\d+)?", s)
    return float(m.group()) if m else None


def clean_unit(v):
    s = (str(v).strip().upper() if v else "")
    if s.startswith("KG"):
        return "KG"
    return "MTR"  # MTR / MTR. / . / blank all -> MTR


def first_token(v):
    """Vendor/cut-master cells hold combos like 'MANTU, VINEY' or 'A / B' — keep the first."""
    if not v:
        return None
    s = str(v).replace("\n", " ").strip()
    s = re.split(r"[,/]| AND ", s, flags=re.IGNORECASE)[0].strip()
    return re.sub(r"\s+", " ", s) or None


def iso(v):
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    return None


def title(s):
    return re.sub(r"\s+", " ", str(s).strip()).title() if s else s


# Hand-authored size breakups for the hero job cards (pulled from real SI sheets)
HERO_SIZES = {
    "SI -01": {"S": 132, "M": 268, "L": 400, "XL": 400, "2XL": 268, "3XL": 132},
}
# Real dispatch log from the SI-01 sheet (sums to 1600)
HERO_DISPATCH = {
    "SI -01": [("2026-01-08", 575), ("2026-01-10", 492), ("2026-01-18", 240),
               ("2026-01-24", 73), ("2026-01-25", 72), ("2026-02-07", 92), ("2026-03-03", 56)],
}


def size_breakup(si, cut):
    if si in HERO_SIZES:
        return [{"size": k, "qty": v} for k, v in HERO_SIZES[si].items()]
    if not cut or cut <= 0:
        return []
    out, run = [], 0
    for i, (sz, r) in enumerate(SIZE_RATIO):
        q = round(cut * r) if i < len(SIZE_RATIO) - 1 else round(cut) - run
        run += q
        if q > 0:
            out.append({"size": sz, "qty": q})
    return out


def dispatch_events(si, disp, order_date):
    if si in HERO_DISPATCH:
        return [{"date": d, "qty": q} for d, q in HERO_DISPATCH[si]]
    if not disp or disp <= 0:
        return []
    base = order_date or "2026-01-15"
    d0 = datetime.date.fromisoformat(base)
    # deterministic 2-event split
    q1 = round(disp * 0.6)
    return [
        {"date": (d0 + datetime.timedelta(days=6)).isoformat(), "qty": q1},
        {"date": (d0 + datetime.timedelta(days=16)).isoformat(), "qty": round(disp) - q1},
    ]


def si_num(s):
    """Extract the integer SI number from 'SI -01' / 'SI-242' for cross-sheet matching."""
    m = re.search(r"\d+", str(s) if s is not None else "")
    return int(m.group()) if m else None


def fabric_colors(wb, si_sheets, si):
    """Read the per-colour 'FABRIC DETAIL' block from a job card's own SI sheet.

    Each SI sheet has a block headed 'S.NO | FABRIC COLOR | REQ. PCS |
    REQD QTY (IN MTR.) | ROLLS'. ~55 of the sheets fill it in with the real
    per-colour fabric issued (the rest are blank templates). Returns
    [{color, reqPcs, reqMtr, rolls}] for the filled colour lines, else [].
    """
    sheet = si_sheets.get(si_num(si))
    if not sheet:
        return []
    ws = wb[sheet]
    grid = list(ws.iter_rows(min_row=1, max_row=22, values_only=True))
    # Locate the 'FABRIC COLOR' header cell -> (row index, colour column index).
    hr = hc = None
    for ri, row in enumerate(grid):
        for ci, v in enumerate(row or ()):
            if isinstance(v, str) and "FABRIC COLOR" in v.upper():
                hr, hc = ri, ci
                break
        if hr is not None:
            break
    if hr is None:
        return []
    out = []
    for row in grid[hr + 1:hr + 11]:  # up to 8 colour lines + slack
        if not row or hc >= len(row):
            continue
        raw = row[hc]
        color = re.sub(r"\s+", " ", str(raw).strip()).upper() if raw not in (None, "") else ""
        if not color or color in ("TTL", "TOTAL"):
            continue
        get = lambda off: num(row[hc + off]) if hc + off < len(row) else None
        req_pcs, req_mtr, rolls = get(1), get(2), get(3)
        if req_pcs is None and req_mtr is None and rolls is None:
            continue  # colour name with no numbers — skip empty template echo
        out.append({"color": color, "reqPcs": req_pcs, "reqMtr": req_mtr, "rolls": rolls})
    return out


def main():
    wb = openpyxl.load_workbook(WB, read_only=True, data_only=True)
    si_sheets = {si_num(n): n for n in wb.sheetnames if n.upper().startswith("SI-")}

    vendors = {}
    for r in wb["VENDOR_LIST"].iter_rows(values_only=True):
        if r and r[0]:
            nm = title(r[0])
            kind = "INHOUSE" if re.search(r"IN\s?HOUSE", str(r[0]), re.I) else "EXTERNAL"
            vendors[nm] = {"name": nm, "kind": kind}

    fabrics = {}      # name -> {name, unit, consumed_sum}
    cmasters = set()
    styles = {}       # styleNo -> style dict
    jobcards = []

    rows = list(wb["CONSOLIDATED STATUS"].iter_rows(values_only=True))[1:]
    si_counter = 0
    for r in rows:
        if not any(r):
            continue
        si = str(r[0]).strip() if r[0] else None
        if not si:
            continue
        si_counter += 1

        item = title(r[C_ITEM]) or "Unknown Item"
        sku = (str(r[C_SKU]).strip() if r[C_SKU] else None)
        if not sku or any(j in sku.upper() for j in JUNK):
            sku = f"AUTO-{si.replace(' ', '')}-{si_counter}"
        unit = clean_unit(r[C_UNIT])
        avg = num(r[C_AVG])
        mrp = num(r[C_MRP])
        fabric_name = title(r[C_FABRIC]) if r[C_FABRIC] else None

        if fabric_name:
            f = fabrics.setdefault(fabric_name, {"name": fabric_name, "unit": unit, "consumed": 0.0})
            consumed = num(r[C_CONSUMED])
            cut_v, _ = num(r[C_CUT]), None
            if consumed is None and cut_v and avg:
                consumed = cut_v * avg
            f["consumed"] += (consumed or 0.0)

        if sku not in styles:
            styles[sku] = {
                "styleNo": sku, "sku": sku, "itemDesc": item, "mrp": mrp,
                "avgConsumption": avg, "unit": unit, "fabric": fabric_name,
                "category": item.split()[0] if item else None,
            }

        vendor_name = first_token(r[C_VENDOR])
        if vendor_name:
            vendor_name = title(vendor_name)
            vendors.setdefault(vendor_name, {"name": vendor_name, "kind":
                               "INHOUSE" if "Inhouse" in vendor_name or "In House" in vendor_name else "EXTERNAL"})
        cm = first_token(r[C_CUTMASTER])
        if cm:
            cm = title(cm)
            cmasters.add(cm)

        cut = num(r[C_CUT]) or 0.0
        disp = num(r[C_DISP]) or 0.0
        order_date = iso(r[C_ORDER_DATE])
        status_raw = (str(r[C_STATUS]).strip().upper() if r[C_STATUS] else "")
        bal = cut - disp
        status = "CLOSED" if ("CLOSE" in status_raw or (cut > 0 and bal <= 0 and disp > 0)) else "ACTIVE"

        jobcards.append({
            "siNo": si,
            "orderDate": order_date,
            "styleNo": sku,
            "vendor": vendor_name,
            "cuttingMaster": cm,
            "cutQty": cut,
            "dispatchedQty": min(disp, cut) if cut else disp,
            "avgConsumption": avg,
            "fabricIssued": num(r[C_FAB_ISSUED]),
            "fabricConsumed": num(r[C_CONSUMED]),
            "fabricIssueDate": iso(r[C_FAB_ISSUE_DATE]),
            "cuttingIssuedOn": iso(r[C_CUT_ISSUED_ON]),
            "plannedEtd": iso(r[C_ETD]),
            "status": status,
            "remark": (str(r[C_REMARK]).strip() if r[C_REMARK] else None),
            "fabric": fabric_name,
            "fabricColors": fabric_colors(wb, si_sheets, si),
            "sizeBreakup": size_breakup(si, cut),
            "dispatches": dispatch_events(si, min(disp, cut) if cut else disp, order_date),
        })

    # synthesize opening stock with a per-fabric buffer so live inventory shows a
    # realistic spread of OK / low / short states (instead of a flat 87% everywhere).
    BUFFERS = [0.88, 1.04, 1.22, 1.55, 1.95, 2.5]
    HERO_BUFFER = {"Ns Cotton": 1.7, "Max Polo": 1.5, "Playcool Eco": 1.45, "Fitness": 1.35}
    fabric_list = []
    for nm, f in fabrics.items():
        if nm in HERO_BUFFER:
            mult = HERO_BUFFER[nm]
        else:
            mult = BUFFERS[sum(ord(c) for c in nm) % len(BUFFERS)]
        opening = round(max(f["consumed"] * mult, 500), 1)
        fabric_list.append({"name": nm, "unit": f["unit"], "openingStock": opening})

    seed = {
        "vendors": list(vendors.values()),
        "cuttingMasters": [{"name": c} for c in sorted(cmasters)],
        "fabrics": fabric_list,
        "styles": list(styles.values()),
        "jobCards": jobcards,
    }
    with open(OUT, "w") as fh:
        json.dump(seed, fh, indent=1)

    print(f"vendors={len(seed['vendors'])} cuttingMasters={len(seed['cuttingMasters'])} "
          f"fabrics={len(fabric_list)} styles={len(styles)} jobCards={len(jobcards)}")
    tc = sum(j["cutQty"] for j in jobcards)
    td = sum(j["dispatchedQty"] for j in jobcards)
    with_colors = sum(1 for j in jobcards if j.get("fabricColors"))
    color_lines = sum(len(j.get("fabricColors") or []) for j in jobcards)
    print(f"total cut={tc:,.0f} dispatched={td:,.0f} -> seed.json")
    print(f"per-colour fabric: {with_colors} job cards, {color_lines} colour lines extracted")


if __name__ == "__main__":
    main()
