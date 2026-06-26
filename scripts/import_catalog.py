#!/usr/bin/env python3
"""Build seed_catalog.json — the commercial layer for the Sportsun ERP demo.

Pulls three real sources and stitches them together:
  1. Product master  ← sportsun-os/sportsun.sqlite `products` (223 SKUs)
  2. BOM             ← BOM 15-5-2026.xlsx (7 styles)
  3. Trims store     ← TIRMS STOCK FINAL ….csv (1,091 items + movements)

Links resolved here (so prisma/seed.ts stays a dumb inserter):
  • product → workbook Style, by normalized SKU (exact) then 3–4 digit style#
  • BOM     → product, by the 3-digit sheet-name code vs SKU trailing digits
  • BOM line → trim item, fuzzy within the same material family
Also synthesizes a handful of production orders (target = 2 × avg monthly sale).

Run AFTER scripts/import_workbook.py (it reads prisma/seed.json for the style link).
Paths are env-overridable: SPORTSUN_DB / BOM_XLSX / TRIMS_CSV.
"""
import csv
import datetime
import json
import os
import re
import sqlite3
from typing import Any, Optional

import openpyxl

HERE = os.path.dirname(__file__)
SPORTSUN_DB = os.environ.get(
    "SPORTSUN_DB",
    os.path.expanduser("~/Desktop/projects/Sportsun/sportsun-os/sportsun.sqlite"),
)
BOM_XLSX = os.environ.get(
    "BOM_XLSX",
    os.path.expanduser("~/Downloads/Sportsun Factory Data/BOM 15-5-2026.xlsx"),
)
TRIMS_CSV = os.environ.get(
    "TRIMS_CSV",
    os.path.expanduser(
        "~/Downloads/Sportsun Factory Data/TIRMS STOCK FINAL - FINAL STOCK  IS NEW.csv"
    ),
)
NOTION_CSV = os.environ.get(
    "NOTION_CSV",
    os.path.expanduser(
        "~/Desktop/projects/Sportsun/sportsun-os/src/data/notion/products.csv"
    ),
)
SEED_JSON = os.path.join(HERE, "..", "prisma", "seed.json")
OUT = os.path.join(HERE, "..", "prisma", "seed_catalog.json")

# Default size ratio (matches the historical SIZE_RATIO used in the app).
SIZE_RATIO: list[list[Any]] = [
    ["S", 0.08], ["M", 0.17], ["L", 0.25], ["XL", 0.25], ["2XL", 0.17], ["3XL", 0.08]
]
# 1–2 size categories → a custom ratio (the "custom formula" case).
FREE_SIZE_CATEGORIES = {"Accessories", "BAGS", "SHOES"}

# Synthesized per-product color masters (no color data in source). Per head category.
CATEGORY_PALETTE: dict[str, list[str]] = {
    "Polo": ["NAVY", "BLACK", "WHITE", "MAROON", "ROYAL BLUE", "BOTTLE GREEN"],
    "Roundneck": ["BLACK", "WHITE", "GREY MELANGE", "NAVY", "RED", "ROYAL BLUE"],
    "Trackpant": ["BLACK", "NAVY", "CHARCOAL", "GREY"],
    "TRACKSUIT": ["BLACK", "NAVY", "ROYAL BLUE", "GREY", "RED"],
    "TRACK UPPER W MESH": ["BLACK", "NAVY", "GREY"],
    "Shorts": ["BLACK", "NAVY", "GREY", "ROYAL BLUE"],
    "Women": ["BLACK", "MAROON", "TEAL", "CORAL", "NAVY"],
    "Kids": ["RED", "ROYAL BLUE", "GREEN", "YELLOW", "BLACK"],
    "Vest / Cut Sleeves": ["BLACK", "WHITE", "GREY", "NAVY"],
    "SPORTS KITS": ["BLUE", "RED", "GREEN", "YELLOW", "WHITE"],
    "Roundneck": ["BLACK", "WHITE", "GREY MELANGE", "NAVY", "RED"],
    "TIGHTS": ["BLACK", "NAVY", "CHARCOAL"],
    "Accessories": ["BLACK", "WHITE"],
}
DEFAULT_PALETTE = ["BLACK", "WHITE", "NAVY", "GREY"]
COLOR_HEX: dict[str, str] = {
    "BLACK": "#1e2330", "WHITE": "#f4f5f7", "NAVY": "#1e293b", "GREY": "#9aa3b2",
    "GREY MELANGE": "#b8bdc7", "CHARCOAL": "#374151", "RED": "#e11d48",
    "MAROON": "#7f1d1d", "ROYAL BLUE": "#2563eb", "BLUE": "#2563eb",
    "BOTTLE GREEN": "#14532d", "GREEN": "#16a34a", "TEAL": "#0d9488",
    "CORAL": "#fb7185", "YELLOW": "#eab308",
}

STATUS_MAP = {
    "ACTIVE": "ACTIVE",
    "NEW ARTICLE": "NEW_ARTICLE",
    "FUTURE PLAN": "FUTURE_PLAN",
    "DISCONTINUED": "DISCONTINUED",
    "IN PROCESS": "IN_PROCESS",
}

# Material/trim family classification (checked in order — specific before generic).
FAMILY_RULES: list[tuple[tuple[str, ...], str]] = [
    (("ZIP",), "ZIP"),
    (("NIWAD",), "NIWAD TAP"),
    (("NADA",), "NADA"),
    (("ELASTIC", "SPAIKA"), "ELASTIC"),
    (("BUTTON",), "BUTTON"),
    (("LABLE", "LABEL"), "LABEL"),
    (("SHOLDER", "SHOULDER"), "SHOULDER TAPE"),
    (("FUSING",), "FUSING"),
    (("HANGER", "LUPPI"), "HANGER"),
    (("STICKER",), "STICKER"),
    (("TICKLE", "TIKLI", "TICKET"), "SIZE TICKET"),
    (("GATTA", "COLLOR", "COLLAR"), "COLLAR"),
    (("HEAT", "D.HEAT", "DHEAT"), "HEAT TRANSFER"),
    (("CONE", "THREAD"), "THREAD"),
    (("POLLY", "POLYBAG", "POLY BAG"), "PACKAGING"),
    (("CARDBOARD", "CARTON", "BOX"), "PACKAGING"),
    (("TAG",), "TAG"),
    (("CORD", "DORI", "NIWAR"), "CORD"),
    (("REFLECT",), "REFLECTOR"),
]

# Plausible avg monthly sale by head category (demo synthesis for production orders).
MONTHLY_SALE = {
    "Polo": 900,
    "Roundneck": 1500,
    "Trackpant": 900,
    "TRACKSUIT": 420,
    "Shorts": 700,
    "Women": 500,
    "Kids": 600,
    "Vest / Cut Sleeves": 800,
    "SPORTS KITS": 300,
    "Accessories": 1200,
}


def num(v: Any) -> Optional[float]:
    """Coerce a cell to float, or None if junk/empty. Keeps negatives."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("₹", "")
    if not s:
        return None
    m = re.search(r"-?\d+(\.\d+)?", s)
    return float(m.group()) if m else None


def normalize(s: Any) -> str:
    """Strip + collapse + uppercase + drop punctuation (the ops-doc matching rule)."""
    if s is None:
        return ""
    return re.sub(r"[^A-Z0-9]", "", str(s).upper())


def tokens(s: Any) -> set[str]:
    if s is None:
        return set()
    return {t for t in re.split(r"[^A-Z0-9]+", str(s).upper()) if len(t) > 1}


def style_number(sku: Any) -> Optional[str]:
    """Trailing 3–4 digit STYLE number; None for 2-digit (too noisy to match on)."""
    m = re.search(r"(\d{3,4})(?:[/-]\d+)?$", str(sku).strip()) if sku else None
    return m.group(1) if m else None


def family_of(text: Any) -> Optional[str]:
    up = str(text or "").upper()
    for keys, fam in FAMILY_RULES:
        if any(k in up for k in keys):
            return fam
    return None


def iso_date(v: Any) -> Optional[str]:
    """Parse messy dd/mm/yyyy | dd-mm-yyyy | datetime → ISO, else None."""
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    s = str(v or "").strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def load_notion() -> dict[str, dict[str, Any]]:
    """Notion products.csv → {normKey: {fabric, imageUrl}} keyed by normSku and normName."""
    out: dict[str, dict[str, Any]] = {}
    if not os.path.exists(NOTION_CSV):
        return out
    with open(NOTION_CSV, newline="", encoding="utf-8-sig", errors="replace") as fh:
        rows = list(csv.reader(fh))
    if not rows:
        return out
    hdr = [h.strip().lower() for h in rows[0]]

    def col(*names: str) -> Optional[int]:
        for n in names:
            if n.lower() in hdr:
                return hdr.index(n.lower())
        return None

    c_name, c_sku = col("name"), col("sku code")
    c_fab = col("fabric name")
    c_img = col("master image url", "product image", "ecommerce images url")
    for r in rows[1:]:
        def cell(i: Optional[int]) -> str:
            return r[i].strip() if i is not None and len(r) > i else ""
        info = {"fabric": cell(c_fab) or None, "imageUrl": cell(c_img) or None}
        if not info["fabric"] and not info["imageUrl"]:
            continue
        for key in (normalize(cell(c_sku)), normalize(cell(c_name))):
            if key:
                out.setdefault(key, info)
    return out


def load_styles() -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]]]:
    """Workbook styles from seed.json + a map normalize(styleNo) → style."""
    styles: list[dict[str, Any]] = []
    if os.path.exists(SEED_JSON):
        with open(SEED_JSON) as fh:
            styles = json.load(fh).get("styles", [])
    by_norm = {normalize(s.get("styleNo") or ""): s for s in styles}
    return styles, by_norm


def synth_colors(product: dict[str, Any]) -> list[dict[str, Any]]:
    """Per-product color master from a category palette (no source color data)."""
    pal = CATEGORY_PALETTE.get(product.get("headCategory") or "", DEFAULT_PALETTE)
    return [
        {"name": c, "hex": COLOR_HEX.get(c), "sortOrder": i}
        for i, c in enumerate(pal)
    ]


def size_ratio_for(product: dict[str, Any]) -> list[list[Any]]:
    if (product.get("headCategory") or "") in FREE_SIZE_CATEGORIES:
        return [["FREE", 1.0]]
    return SIZE_RATIO


def load_products() -> list[dict[str, Any]]:
    con = sqlite3.connect(f"file:{SPORTSUN_DB}?mode=ro", uri=True)
    con.row_factory = sqlite3.Row
    rows = con.execute(
        "SELECT id, sku_code, name, head_category, mrp, custom_ws_rate, status, style_group "
        "FROM products"
    ).fetchall()
    con.close()
    out: list[dict[str, Any]] = []
    for r in rows:
        sku = (r["sku_code"] or "").strip()
        cat = (r["head_category"] or "").strip() or None
        status = STATUS_MAP.get((r["status"] or "").strip().upper(), "ACTIVE")
        sno = style_number(sku)
        out.append(
            {
                "extId": r["id"],
                "skuCode": sku,
                "normSku": normalize(sku),
                "styleNo": sno,
                "name": (r["name"] or "").strip(),
                "headCategory": cat,
                "mrp": num(r["mrp"]),
                "customWsRate": num(r["custom_ws_rate"]),
                "status": status,
                "styleGroup": (r["style_group"] or "").strip() or None,
                "bomCode": sno,
            }
        )
    return out


def link_products_to_styles(
    products: list[dict[str, Any]], style_by_norm: dict[str, dict[str, Any]]
) -> None:
    """Set linkedStyleNo + copy production fields (fabric/avg/unit/itemDesc) from the
    matched workbook style. Match by normalized SKU (exact) then 3–4 digit style#."""
    by_norm: dict[str, dict[str, Any]] = dict(style_by_norm)
    by_num: dict[str, dict[str, Any]] = {}
    for snorm, s in style_by_norm.items():
        n = style_number(s.get("styleNo") or "")
        if n:
            by_num.setdefault(n, s)
    exact = fallback = 0
    for p in products:
        st = by_norm.get(p["normSku"])
        if st:
            exact += 1
        elif p["styleNo"] and p["styleNo"] in by_num:
            st = by_num[p["styleNo"]]
            fallback += 1
        p["linkedStyleNo"] = (st or {}).get("styleNo")
        if st:  # copy production-master fields
            p["itemDesc"] = st.get("itemDesc")
            p["avgConsumption"] = st.get("avgConsumption")
            p["unit"] = st.get("unit") or "MTR"
            p["fabric"] = st.get("fabric")
    print(f"  product→style link: {exact} exact + {fallback} fallback = {exact + fallback}")


def unify_master(
    products: list[dict[str, Any]],
    styles: list[dict[str, Any]],
    notion: dict[str, dict[str, Any]],
) -> dict[str, str]:
    """Enrich products (notion fabric/photo, colors, ratios), synthesize Products for
    workbook styles unmatched by any catalog product, and return normSku→productExtId
    covering EVERY style (so every job card maps to a product)."""
    norm_to_ext: dict[str, str] = {}
    for p in products:
        # notion fabric/photo (fabric only if the style didn't already provide one)
        info = notion.get(p["normSku"]) or notion.get(normalize(p["name"]))
        if info:
            p.setdefault("fabric", None)
            if not p.get("fabric"):
                p["fabric"] = info.get("fabric")
            p["imageUrl"] = info.get("imageUrl")
        p.setdefault("itemDesc", p.get("name"))
        p.setdefault("avgConsumption", None)
        p.setdefault("unit", "MTR")
        p.setdefault("fabric", None)
        p.setdefault("imageUrl", None)
        p["colors"] = synth_colors(p)
        p["sizeRatioJson"] = json.dumps(size_ratio_for(p))
        p["colorRatioJson"] = json.dumps(
            [[c["name"], round(1 / max(1, len(p["colors"])), 4)] for c in p["colors"]]
        )
        norm_to_ext[p["normSku"]] = p["extId"]
        if p.get("linkedStyleNo"):
            norm_to_ext.setdefault(normalize(p["linkedStyleNo"]), p["extId"])

    # synthesize a Product for every workbook style not yet covered
    made = 0
    for s in styles:
        snorm = normalize(s.get("styleNo") or "")
        if not snorm or snorm in norm_to_ext:
            continue
        ext = f"STY-{s['styleNo']}"
        cat = s.get("category")
        prod = {
            "extId": ext,
            "skuCode": s["styleNo"],
            "normSku": snorm,
            "styleNo": style_number(s["styleNo"]),
            "name": s.get("itemDesc") or s["styleNo"],
            "itemDesc": s.get("itemDesc"),
            "headCategory": cat,
            "mrp": s.get("mrp"),
            "customWsRate": None,
            "status": "ACTIVE",
            "styleGroup": None,
            "bomCode": None,
            "avgConsumption": s.get("avgConsumption"),
            "unit": s.get("unit") or "MTR",
            "fabric": s.get("fabric"),
            "imageUrl": None,
            "linkedStyleNo": s["styleNo"],
        }
        prod["colors"] = synth_colors(prod)
        prod["sizeRatioJson"] = json.dumps(size_ratio_for(prod))
        prod["colorRatioJson"] = json.dumps(
            [[c["name"], round(1 / max(1, len(prod["colors"])), 4)] for c in prod["colors"]]
        )
        products.append(prod)
        norm_to_ext[snorm] = ext
        made += 1

    style_norm_to_ext = {
        normalize(s.get("styleNo") or ""): norm_to_ext[normalize(s.get("styleNo") or "")]
        for s in styles
        if normalize(s.get("styleNo") or "") in norm_to_ext
    }
    missing = [s["styleNo"] for s in styles if normalize(s.get("styleNo") or "") not in norm_to_ext]
    assert not missing, f"styles with no product: {missing[:10]}"
    print(f"  unified master: +{made} STY products → {len(products)} total; every style mapped")
    return style_norm_to_ext


def synth_users(styles_vendors: list[str]) -> list[dict[str, Any]]:
    """Demo users. The VENDOR user's vendorName must match a real Vendor.name."""
    fashion = next((v for v in styles_vendors if "FASHION" in v.upper()), "Fashion 11")
    return [
        {"username": "admin", "displayName": "Admin", "role": "ADMIN", "vendorName": None},
        {"username": "jyotika", "displayName": "Jyotika", "role": "STAFF", "vendorName": None},
        {"username": "fashion11", "displayName": "Fashion Eleven", "role": "VENDOR", "vendorName": fashion},
        {"username": "satya", "displayName": "Satya", "role": "TRIMS", "vendorName": None},
    ]


def load_boms(products: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_code: dict[str, str] = {}
    for p in products:  # first product per 3-digit code
        if p["bomCode"]:
            by_code.setdefault(p["bomCode"], p["extId"])

    wb = openpyxl.load_workbook(BOM_XLSX, read_only=True, data_only=True)
    boms: list[dict[str, Any]] = []
    for sheet in wb.sheetnames:
        code_m = re.search(r"(\d{3,4})", sheet)  # trust the sheet name, not row 1
        if not code_m:
            continue
        code = code_m.group(1)
        style_name = re.sub(r"^\d{3,4}[\s\-]*", "", sheet).strip() or sheet
        ws = wb[sheet]
        lines: list[dict[str, Any]] = []
        cur_material: Optional[str] = None
        cur_sno: Optional[int] = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 3:
                continue
            cells = list(row)
            joined = " ".join(str(c) for c in cells if c is not None).upper()
            if "BASE COL" in joined:  # secondary cross-table sentinel (301-HS) — stop
                break
            sno_cell = cells[0] if len(cells) > 0 else None
            mat_cell = cells[1] if len(cells) > 1 else None
            col_cell = cells[2] if len(cells) > 2 else None
            qty_cell = cells[3] if len(cells) > 3 else None
            avg_cell = cells[4] if len(cells) > 4 else None
            if mat_cell is not None and str(mat_cell).strip():
                cur_material = str(mat_cell).strip()
                sn = num(sno_cell)
                cur_sno = int(sn) if sn is not None else None
            if cur_material is None:
                continue
            color = str(col_cell).strip() if col_cell is not None and str(col_cell).strip() else None
            avg = str(avg_cell).strip() if avg_cell is not None and str(avg_cell).strip() else None
            if color is None and qty_cell is None and avg is None and mat_cell is None:
                continue
            lines.append(
                {
                    "sNo": cur_sno,
                    "material": cur_material,
                    "color": color,
                    "qty": num(qty_cell),
                    "avg": avg,
                }
            )
        boms.append(
            {
                "code": code,
                "styleName": style_name,
                "productExtId": by_code.get(code),
                "lines": lines,
            }
        )
    return boms


def load_trims() -> dict[str, Any]:
    with open(TRIMS_CSV, newline="", encoding="utf-8-sig", errors="replace") as fh:
        rows = list(csv.reader(fh))
    items: dict[str, dict[str, Any]] = {}  # normName → item (last-wins on dup name)
    for r in rows[2:]:
        if len(r) > 13 and r[13].strip():
            name = r[13].strip()
            items[normalize(name)] = {
                "sno": r[12].strip() if len(r) > 12 and r[12].strip() else None,
                "name": name,
                "normName": normalize(name),
                "family": family_of(name),
                "currentStock": num(r[14]) if len(r) > 14 else None,
                "openingStock": num(r[15]) if len(r) > 15 else None,
            }
    trim_list = list(items.values())
    valid = set(items.keys())

    movements: list[dict[str, Any]] = []
    for r in rows[2:]:
        # STOCK IN (cols 0-5) → RECEIPT
        if len(r) > 3 and r[2].strip() and num(r[3]) is not None:
            n = normalize(r[2])
            if n in valid:
                movements.append(
                    {
                        "type": "RECEIPT",
                        "itemNorm": n,
                        "date": iso_date(r[0]),
                        "invoice": r[1].strip() or None,
                        "qty": num(r[3]),
                        "rate": num(r[4]) if len(r) > 4 else None,
                        "vendor": r[5].strip() or None if len(r) > 5 else None,
                    }
                )
        # STOCK OUT (cols 6-11) → ISSUE
        if len(r) > 9 and r[8].strip() and num(r[9]) is not None:
            n = normalize(r[8])
            if n in valid:
                movements.append(
                    {
                        "type": "ISSUE",
                        "itemNorm": n,
                        "date": iso_date(r[6]),
                        "invoice": r[7].strip() or None,
                        "qty": num(r[9]),
                        "rate": num(r[10]) if len(r) > 10 else None,
                        "vendor": r[11].strip() or None if len(r) > 11 else None,
                    }
                )
    return {"items": trim_list, "movements": movements}


def match_bom_to_trims(boms: list[dict[str, Any]], trims: list[dict[str, Any]]) -> int:
    """Fuzzy-match each BOM line to a trim item within the same family."""
    by_family: dict[str, list[dict[str, Any]]] = {}
    for t in trims:
        if t["family"]:
            by_family.setdefault(t["family"], []).append(t)

    matched = 0
    for bom in boms:
        for line in bom["lines"]:
            fam = family_of(line["material"])
            cands = by_family.get(fam or "", [])
            if not cands:
                line["trimMatchNorm"] = None
                continue
            qtok = tokens(line["color"]) or tokens(line["material"])
            best, best_score = None, -1.0
            for t in cands:
                ttok = tokens(t["name"])
                inter = len(qtok & ttok)
                union = len(qtok | ttok) or 1
                score = inter / union
                if not qtok:  # no colour to match → prefer the highest-stock variant
                    score = (t["currentStock"] or 0) / 1e9
                if score > best_score:
                    best, best_score = t, score
            line["trimMatchNorm"] = best["normName"] if best else None
            if best:
                matched += 1
    return matched


def synth_production_orders(products: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """~14 illustrative orders against real active SKUs; target = 2 × monthly sale."""
    active = [p for p in products if p["status"] in ("ACTIVE", "NEW_ARTICLE")]
    # deterministic spread: prefer linked (recognizable) SKUs first
    active.sort(key=lambda p: (p.get("linkedStyleNo") is None, p["extId"]))
    statuses = ["IN_PRODUCTION", "ORDER_GIVEN", "COMPLETED"]
    urgencies = ["V URGENT", "URGENT", "MODERATE"]
    base = datetime.date(2026, 6, 1)
    orders: list[dict[str, Any]] = []
    for i, p in enumerate(active[:14]):
        sale = MONTHLY_SALE.get(p["headCategory"] or "", 600) + (i % 5) * 40
        orders.append(
            {
                "orderNo": f"PO-{i + 1:02d}",
                "productExtId": p["extId"],
                "orderDate": (base - datetime.timedelta(days=i * 2)).isoformat(),
                "avgMonthlySale": float(sale),
                "targetQty": float(round(2 * sale)),
                "status": statuses[i % 3],
                "urgency": urgencies[i % 3],
                "remarks": None,
            }
        )
    return orders


def load_seed_vendors() -> list[str]:
    if not os.path.exists(SEED_JSON):
        return []
    with open(SEED_JSON) as fh:
        return [v.get("name", "") for v in json.load(fh).get("vendors", [])]


def main() -> None:
    styles, style_by_norm = load_styles()
    notion = load_notion()
    products = load_products()
    link_products_to_styles(products, style_by_norm)
    style_norm_to_ext = unify_master(products, styles, notion)  # enriches + adds STY products
    boms = load_boms(products)
    trims_bundle = load_trims()
    trim_items = trims_bundle["items"]
    trim_moves = trims_bundle["movements"]
    matched = match_bom_to_trims(boms, trim_items)
    orders = synth_production_orders(products)
    users = synth_users(load_seed_vendors())

    seed = {
        "products": products,
        "boms": boms,
        "trimItems": trim_items,
        "trimMovements": trim_moves,
        "productionOrders": orders,
        "users": users,
        "styleNormToProductExtId": style_norm_to_ext,
    }
    with open(OUT, "w") as fh:
        json.dump(seed, fh, indent=1)

    bom_lines = sum(len(b["lines"]) for b in boms)
    linked_boms = sum(1 for b in boms if b["productExtId"])
    with_fabric = sum(1 for p in products if p.get("fabric"))
    with_img = sum(1 for p in products if p.get("imageUrl"))
    print(
        f"products={len(products)} (fabric {with_fabric}, photo {with_img}) "
        f"boms={len(boms)} (linked {linked_boms}, {bom_lines} lines, {matched} trim-matched) "
        f"trimItems={len(trim_items)} trimMovements={len(trim_moves)} "
        f"productionOrders={len(orders)} users={len(users)} -> seed_catalog.json"
    )


if __name__ == "__main__":
    main()
